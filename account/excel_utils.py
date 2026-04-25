# account/excel_utils.py

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone


class ExcelExportUtil:
    """Utility class for exporting visitor data to Excel"""
    
    def __init__(self, visitors, start_date, end_date, status_filter=''):
        self.visitors = visitors
        self.start_date = start_date
        self.end_date = end_date
        self.status_filter = status_filter
        self.wb = Workbook()
        
        # Define styles
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.subheader_font = Font(bold=True, size=10)
        self.header_fill_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_fill_green = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        self.header_fill_orange = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.left_alignment = Alignment(horizontal="left", vertical="center")
    
    def generate(self):
        """Generate complete Excel workbook"""
        self.create_summary_sheet()
        self.create_visitors_sheet()
        self.create_sections_sheet()
        self.create_timeline_sheet()
        return self.wb
    
    def create_summary_sheet(self):
        """Create summary statistics sheet"""
        ws = self.wb.active
        ws.title = "Summary"
        
        # Summary data
        summary_data = [
            ("VISITOR MANAGEMENT SYSTEM - EXPORT REPORT", ""),
            ("", ""),
            ("Generated On", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Date Range", f"{self.start_date} to {self.end_date}"),
            ("Status Filter", self.status_filter if self.status_filter else "All"),
            ("", ""),
            ("OVERALL STATISTICS", ""),
            ("Total Visitors", self.visitors.count()),
            ("", ""),
            ("STATUS BREAKDOWN", ""),
            ("Pending", self.visitors.filter(status='pending').count()),
            ("Partially Approved", self.visitors.filter(status='partially_approved').count()),
            ("Approved", self.visitors.filter(status='approved').count()),
            ("Checked In", self.visitors.filter(status='checked_in').count()),
            ("Checked Out", self.visitors.filter(status='checked_out').count()),
            ("Rejected", self.visitors.filter(status='rejected').count()),
            ("No Show", self.visitors.filter(status='no_show').count()),
            ("Cancelled", self.visitors.filter(status='cancelled').count()),
            ("", ""),
            ("TIME METRICS", ""),
            ("Total Check-ins", self.visitors.filter(actual_check_in__isnull=False).count()),
            ("Total Check-outs", self.visitors.filter(actual_check_out__isnull=False).count()),
            ("Average Visit Duration (min)", self._calculate_avg_duration()),
            ("Total Overtime (min)", self._calculate_total_overtime()),
            ("On-time Arrivals", self._calculate_on_time_arrivals()),
            ("Late Arrivals", self._calculate_late_arrivals()),
            ("Early Arrivals", self._calculate_early_arrivals()),
            ("", ""),
            ("SECTION METRICS", ""),
            ("Total Section Visits", self._calculate_total_section_visits()),
            ("Average Sections per Visitor", self._calculate_avg_sections()),
            ("Total Time in Sections (min)", self._calculate_total_section_time()),
        ]
        
        row = 1
        for item in summary_data:
            label, value = item
            
            # Apply styles
            cell_label = ws.cell(row=row, column=1, value=label)
            cell_value = ws.cell(row=row, column=2, value=value)
            
            # Format headers
            if label in ["OVERALL STATISTICS", "STATUS BREAKDOWN", "TIME METRICS", "SECTION METRICS"]:
                cell_label.font = self.subheader_font
                cell_label.fill = self.header_fill_blue
                cell_label.alignment = self.center_alignment
            elif label and not value:  # Empty row
                pass
            else:
                cell_label.font = Font(bold=True)
                cell_label.alignment = self.left_alignment
            
            cell_value.alignment = self.left_alignment
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
    
    def create_visitors_sheet(self):
        """Create main visitors data sheet"""
        ws = self.wb.create_sheet("Visitors")
        
        # Define headers
        headers = [
            "ID", "Visitor Name", "Email", "Phone", "Company", "Purpose",
            "Site", "Status", "Designated Check-in", "Designated Check-out",
            "Actual Check-in", "Actual Check-out", "Early (min)", "Late (min)",
            "Duration (min)", "Overtime (min)", "Created By", "Approvers",
            "Vehicle Number", "ID Card", "Host Department", "Meeting Room",
            "Sections Visited", "Total Section Time (min)", "Created At", "Updated At"
        ]
        
        # Apply header styles
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill_blue
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Add data rows
        for row_idx, visitor in enumerate(self.visitors, start=2):
            self._add_visitor_row(ws, row_idx, visitor)
        
        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _add_visitor_row(self, ws, row_idx, visitor):
        """Add a single visitor row to the sheet"""
        
        # Basic Info
        ws.cell(row=row_idx, column=1, value=visitor.id)
        ws.cell(row=row_idx, column=2, value=visitor.full_name or '')
        ws.cell(row=row_idx, column=3, value=visitor.email or '')
        ws.cell(row=row_idx, column=4, value=visitor.phone_number or '')
        ws.cell(row=row_idx, column=5, value=visitor.company_name or '')
        ws.cell(row=row_idx, column=6, value=visitor.purpose_of_visit or '')
        ws.cell(row=row_idx, column=7, value=visitor.site.name if visitor.site else '')
        ws.cell(row=row_idx, column=8, value=visitor.status or '')
        
        # Times
        ws.cell(row=row_idx, column=9, value=self._format_datetime(visitor.designated_check_in))
        ws.cell(row=row_idx, column=10, value=self._format_datetime(visitor.designated_check_out))
        ws.cell(row=row_idx, column=11, value=self._format_datetime(visitor.actual_check_in))
        ws.cell(row=row_idx, column=12, value=self._format_datetime(visitor.actual_check_out))
        
        # Metrics
        ws.cell(row=row_idx, column=13, value=visitor.early_arrival_minutes or 0)
        ws.cell(row=row_idx, column=14, value=visitor.late_arrival_minutes or 0)
        ws.cell(row=row_idx, column=15, value=visitor.visit_duration_minutes or 0)
        ws.cell(row=row_idx, column=16, value=visitor.overtime_minutes or 0)
        
        # Staff Info
        ws.cell(row=row_idx, column=17, value=visitor.created_by.full_name if visitor.created_by else '')
        ws.cell(row=row_idx, column=18, value=self._get_approver_names(visitor))
        
        # Additional Info
        ws.cell(row=row_idx, column=19, value=visitor.vehicle_number or '')
        ws.cell(row=row_idx, column=20, value=visitor.id_card_number or '')
        ws.cell(row=row_idx, column=21, value=visitor.host_department or '')
        ws.cell(row=row_idx, column=22, value=visitor.meeting_room or '')
        
        # Section Stats
        ws.cell(row=row_idx, column=23, value=self._get_completed_sections_count(visitor))
        ws.cell(row=row_idx, column=24, value=self._get_total_section_time(visitor))
        
        # Timestamps
        ws.cell(row=row_idx, column=25, value=self._format_datetime(visitor.created_at))
        ws.cell(row=row_idx, column=26, value=self._format_datetime(visitor.updated_at))
        
        # Apply borders and alignment
        for col_idx in range(1, 27):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = self.border
            if col_idx in [2, 3, 4, 5, 6, 17, 18]:  # Text columns
                cell.alignment = self.left_alignment
            else:
                cell.alignment = self.center_alignment
    
    def create_sections_sheet(self):
        """Create detailed section tracking sheet"""
        ws = self.wb.create_sheet("Section Details")
        
        headers = [
            "Visitor ID", "Visitor Name", "Visitor Email", "Visitor Phone",
            "Section Name", "Section Type", "Location", "Requires Escort",
            "Check-in Time", "Check-out Time", "Duration (minutes)",
            "Status", "Checked In By", "Checked Out By"
        ]
        
        # Apply header styles
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill_green
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        row_idx = 2
        for visitor in self.visitors:
            for tracking in visitor.section_trackings.all():
                self._add_section_row(ws, row_idx, visitor, tracking)
                row_idx += 1
        
        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        
        ws.freeze_panes = 'A2'
    
    def _add_section_row(self, ws, row_idx, visitor, tracking):
        """Add a single section tracking row"""
        
        ws.cell(row=row_idx, column=1, value=visitor.id)
        ws.cell(row=row_idx, column=2, value=visitor.full_name)
        ws.cell(row=row_idx, column=3, value=visitor.email)
        ws.cell(row=row_idx, column=4, value=visitor.phone_number)
        ws.cell(row=row_idx, column=5, value=tracking.section.name)
        ws.cell(row=row_idx, column=6, value=tracking.section.get_section_type_display())
        ws.cell(row=row_idx, column=7, value=tracking.section.location.name if tracking.section.location else '')
        ws.cell(row=row_idx, column=8, value="Yes" if tracking.section.requires_escort else "No")
        ws.cell(row=row_idx, column=9, value=self._format_datetime(tracking.section_check_in))
        ws.cell(row=row_idx, column=10, value=self._format_datetime(tracking.section_check_out))
        ws.cell(row=row_idx, column=11, value=tracking.duration_minutes)
        ws.cell(row=row_idx, column=12, value=tracking.get_status_display())
        ws.cell(row=row_idx, column=13, value=tracking.checked_in_by.full_name if tracking.checked_in_by else '')
        ws.cell(row=row_idx, column=14, value=tracking.checked_out_by.full_name if tracking.checked_out_by else '')
        
        # Apply borders and alignment
        for col_idx in range(1, 15):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = self.border
            cell.alignment = self.center_alignment
    
    def create_timeline_sheet(self):
        """Create daily timeline sheet"""
        ws = self.wb.create_sheet("Daily Timeline")
        
        headers = ["Date", "Total Visitors", "Checked In", "Checked Out", 
                   "Pending", "Approved", "Rejected", "No Show"]
        
        # Apply header styles
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill_orange
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Group by date
        from collections import defaultdict
        daily_stats = defaultdict(lambda: {
            'total': 0, 'checked_in': 0, 'checked_out': 0, 
            'pending': 0, 'approved': 0, 'rejected': 0, 'no_show': 0
        })
        
        for visitor in self.visitors:
            date_key = visitor.created_at.date().isoformat()
            daily_stats[date_key]['total'] += 1
            status = visitor.status
            
            if status == 'checked_in':
                daily_stats[date_key]['checked_in'] += 1
            elif status == 'checked_out':
                daily_stats[date_key]['checked_out'] += 1
            elif status == 'pending':
                daily_stats[date_key]['pending'] += 1
            elif status == 'approved':
                daily_stats[date_key]['approved'] += 1
            elif status == 'rejected':
                daily_stats[date_key]['rejected'] += 1
            elif status == 'no_show':
                daily_stats[date_key]['no_show'] += 1
        
        # Add data rows
        row_idx = 2
        for date, stats in sorted(daily_stats.items()):
            ws.cell(row=row_idx, column=1, value=date)
            ws.cell(row=row_idx, column=2, value=stats['total'])
            ws.cell(row=row_idx, column=3, value=stats['checked_in'])
            ws.cell(row=row_idx, column=4, value=stats['checked_out'])
            ws.cell(row=row_idx, column=5, value=stats['pending'])
            ws.cell(row=row_idx, column=6, value=stats['approved'])
            ws.cell(row=row_idx, column=7, value=stats['rejected'])
            ws.cell(row=row_idx, column=8, value=stats['no_show'])
            
            # Apply borders and alignment
            for col_idx in range(1, 9):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = self.border
                cell.alignment = self.center_alignment
            
            row_idx += 1
        
        ws.column_dimensions['A'].width = 15
        for col_idx in range(2, 9):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        ws.freeze_panes = 'A2'
    
    # Helper methods
    def _format_datetime(self, dt):
        """Format datetime for Excel display"""
        if dt:
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        return ''
    
    def _get_approver_names(self, visitor):
        """Get comma-separated list of approver names"""
        return ', '.join([approver.full_name for approver in visitor.selected_approvers.all()])
    
    def _get_completed_sections_count(self, visitor):
        """Get count of completed sections"""
        return visitor.section_trackings.filter(status='completed').count()
    
    def _get_total_section_time(self, visitor):
        """Get total time spent in sections"""
        return sum([t.duration_minutes for t in visitor.section_trackings.all()])
    
    def _calculate_avg_duration(self):
        """Calculate average visit duration"""
        durations = [v.visit_duration_minutes for v in self.visitors if v.visit_duration_minutes > 0]
        return round(sum(durations) / len(durations), 2) if durations else 0
    
    def _calculate_total_overtime(self):
        """Calculate total overtime minutes"""
        return sum([v.overtime_minutes for v in self.visitors])
    
    def _calculate_on_time_arrivals(self):
        """Count on-time arrivals (within 5 minutes of designated time)"""
        on_time = 0
        for v in self.visitors:
            if v.actual_check_in and v.designated_check_in:
                diff = abs((v.actual_check_in - v.designated_check_in).total_seconds() / 60)
                if diff <= 5:
                    on_time += 1
        return on_time
    
    def _calculate_late_arrivals(self):
        """Count late arrivals"""
        return self.visitors.filter(late_arrival_minutes__gt=0).count()
    
    def _calculate_early_arrivals(self):
        """Count early arrivals"""
        return self.visitors.filter(early_arrival_minutes__gt=0).count()
    
    def _calculate_total_section_visits(self):
        """Calculate total section visits"""
        return sum([self._get_completed_sections_count(v) for v in self.visitors])
    
    def _calculate_avg_sections(self):
        """Calculate average sections visited per visitor"""
        total_sections = self._calculate_total_section_visits()
        return round(total_sections / self.visitors.count(), 2) if self.visitors.count() > 0 else 0
    
    def _calculate_total_section_time(self):
        """Calculate total time spent in all sections"""
        return sum([self._get_total_section_time(v) for v in self.visitors])